"""
File Artifact Example - Model Saves CSV → Script Converts to .xlsx

WHAT THIS DEMONSTRATES:
- Model saves structured CSV files to artifacts_dir (basic files it can produce)
- A post-generation script converts those CSVs to a formatted .xlsx
- Clean separation: model produces data files, code packages them
- The final deliverable is a file the model could never produce directly

WHY THIS MATTERS:
- Models can write text files. Stakeholders need Excel, Word, PDF, PPT.
- By having the model save basic files (CSV, MD, JSON) as artifacts,
  a deterministic script can convert to ANY target format.
- This extends to 100s of file types without teaching the model anything new.

FLOW:
┌──────────────────────────────────────────────────────────┐
│  HARNESS (with code execution + artifacts_dir)           │
│  → Reads vendor quotes (attached file)                   │
│  → Analyzes data, saves CSV files to artifacts_dir       │
└──────────────────────────────────────────────────────────┘
                          ↓  artifacts_dir contains .csv files
┌──────────────────────────────────────────────────────────┐
│  POST-GENERATION SCRIPT                                  │
│  → Reads CSV artifacts from disk                         │
│  → Writes formatted .xlsx with multiple sheets           │
│  → Final artifact: a rich file from basic model output   │
└──────────────────────────────────────────────────────────┘

ALTERNATIVE: MODEL RUNS THE CONVERSION ITSELF
Instead of a post-generation script, the model can execute a pre-built converter
directly via bash/execute_code — no imports, no documentation needed.

  1. Write a CLI converter: utils/csv_to_xlsx.py artifacts/ report.xlsx
  2. Tell the model: "run utils/csv_to_xlsx.py <csv_dir> <output>" after saving CSVs
  3. Model just executes it like any shell command

The model doesn't write conversion code — it runs yours. Same way it would
call `python utils/weather.py London --forecast`. Execution logs give you
a full audit trail of what ran with what args.

EXTRA SETUP NEEDED:
- .env file with API keys
- uv pip install openpyxl
"""

import csv
import sys
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

from verif import RLHarness, HistoryEntry, Attachment, Prompt
from verif.config import ProviderConfig
from verif.executor import SubprocessExecutor

# =============================================================================
# EVENT HANDLER
# =============================================================================

def on_event(event: HistoryEntry):
    if event.entry_type == "tool_call":
        print(f"  -> {event.content}")

# =============================================================================
# SETUP
# =============================================================================

ARTIFACTS_DIR = Path(__file__).parent / "artifacts"
ARTIFACTS_DIR.mkdir(exist_ok=True)

# Clean previous artifacts
for f in ARTIFACTS_DIR.glob("*.csv"):
    f.unlink()
for f in ARTIFACTS_DIR.glob("*.xlsx"):
    f.unlink()

QUOTES_FILE = Path(__file__).parent.parent / "example_tasks" / "quotes_headlamp.txt"

with open(QUOTES_FILE) as f:
    preview = f.read()

attachment = Attachment(
    content=str(QUOTES_FILE.absolute()),
    mime_type="text/plain",
    name="quotes_headlamp.txt",
    preview=preview,
)

# =============================================================================
# TASK: Model saves CSV files to artifacts_dir
# =============================================================================

TASK: Prompt = [
    f"""Analyze the attached vendor quotations for Model I headlamp and produce
a 4-year NPV comparison. Use a 10% discount rate.
30% of annual volume is top variant, 70% is base variant.

You MUST save your output as CSV files using execute_code. Save each file to
the artifacts directory: {ARTIFACTS_DIR.absolute()}

Save these 4 CSV files (use exact filenames):

1. vendor_comparison.csv
   Columns: vendor,part_price_top,part_price_base,tooling,rd,lead_time_weeks,fx_exposure

2. annual_cost_projection.csv
   Columns: vendor,year,volume,top_units,base_units,part_cost,tooling_amortized,rd_amortized,total_cost

3. npv_summary.csv
   Columns: vendor,year_1_cost,year_2_cost,year_3_cost,year_4_cost,total_npv,rank

4. recommendation.csv
   Columns: factor,autolantic,vendocrat,solimoto,winner
   Factors: total NPV, lead time, FX risk, scale pricing advantage

Use Python csv module to write proper CSV files. Do the actual math.
For Solimoto, use the tiered pricing (< or >= 100000 threshold).
Your text answer should be a brief summary of findings only.
""",
    attachment,
]

# =============================================================================
# CONFIGURATION
# =============================================================================

executor = SubprocessExecutor(
    artifacts_dir=str(ARTIFACTS_DIR),
    timeout=60,
)

harness = RLHarness(
    provider=ProviderConfig(name="gemini", thinking_level="MEDIUM"),
    enable_search=False,
    enable_code=True,
    code_executor=executor,
    artifacts_dir=str(ARTIFACTS_DIR),
    on_event=on_event,
)

# =============================================================================
# STAGE 1: Model generates CSV artifacts
# =============================================================================

print("=" * 60)
print("STAGE 1: Model analyzes data and saves CSV artifacts")
print("=" * 60)

try:
    result = harness.run_single(TASK)
except Exception as e:
    print(f"\nError during execution: {e}", file=sys.stderr)
    sys.exit(1)

if not result.answer:
    print("\nNo answer was generated", file=sys.stderr)
    sys.exit(1)

# Check what CSVs the model created
csv_files = sorted(ARTIFACTS_DIR.glob("*.csv"))
print(f"\nModel created {len(csv_files)} CSV artifact(s):")
for f in csv_files:
    print(f"  {f.name} ({f.stat().st_size:,} bytes)")

if not csv_files:
    print("\nNo CSV artifacts found. Model output:", file=sys.stderr)
    print(result.answer, file=sys.stderr)
    sys.exit(1)

# =============================================================================
# STAGE 2: Convert CSV artifacts to .xlsx
# =============================================================================

print("\n" + "=" * 60)
print("STAGE 2: Converting CSV artifacts to .xlsx")
print("=" * 60)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: uv pip install openpyxl", file=sys.stderr)
    sys.exit(1)

wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

for csv_file in csv_files:
    # Sheet name from filename: vendor_comparison.csv -> Vendor Comparison
    sheet_name = csv_file.stem.replace("_", " ").title()[:31]

    with open(csv_file) as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    if not rows:
        continue

    ws = wb.create_sheet(title=sheet_name)
    headers = list(rows[0].keys())

    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data with numeric conversion
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            val = row.get(h, "")
            cleaned = val.replace(",", "").strip()
            try:
                val = int(cleaned)
            except ValueError:
                try:
                    val = float(cleaned)
                except ValueError:
                    pass
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    print(f"  Sheet: {sheet_name} ({len(rows)} rows)")

output_xlsx = ARTIFACTS_DIR / "vendor_npv_analysis.xlsx"
wb.save(str(output_xlsx))
print(f"\n  -> {output_xlsx.name} ({output_xlsx.stat().st_size:,} bytes)")

# =============================================================================
# SAVE EXECUTION TRACE
# =============================================================================

output_dir = Path(__file__).parent / "outputs"
output_dir.mkdir(parents=True, exist_ok=True)

output_md = output_dir / "file_artifact_output.md"
with open(output_md, "w") as f:
    f.write("# File Artifact Example: CSV Artifacts -> .xlsx\n\n")
    f.write(f"## Task\nVendor NPV analysis from `{QUOTES_FILE.name}`\n\n")
    f.write(f"## Model Summary\n\n{result.answer}\n\n")
    f.write("---\n\n")
    f.write("## CSV Artifacts (model-generated)\n\n")
    for csv_file in csv_files:
        f.write(f"### {csv_file.stem}\n```csv\n{csv_file.read_text()}```\n\n")
    f.write("---\n\n")
    f.write("## Final Artifact (script-converted)\n\n")
    f.write(f"- `{output_xlsx.name}` ({output_xlsx.stat().st_size:,} bytes)\n")
    f.write(f"- Sheets: {', '.join(c.stem.replace('_', ' ').title() for c in csv_files)}\n\n")
    f.write("---\n\n")
    f.write(f"## Rubric\n\n{result.rubric}\n\n")
    f.write("---\n\n")
    f.write("## Execution Trace\n\n")
    f.write(harness.get_history_markdown())

print(f"  Trace: {output_md}")

# =============================================================================
# SUMMARY
# =============================================================================

print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"""
  Model saved: {len(csv_files)} CSV files to artifacts_dir
  Script converted to: {output_xlsx.name}
  Sheets: {[c.stem.replace('_', ' ').title() for c in csv_files]}

  The model wrote basic CSV files (text it's good at).
  The script converted them to a formatted Excel workbook.
  This pattern works for any target format:
    CSV -> .xlsx (openpyxl)
    MD  -> .docx (python-docx)
    MD  -> .pdf  (weasyprint)
    CSV -> .pptx charts (python-pptx)
    JSON -> any API payload
""")
